#!/usr/bin/env python3
"""
Zaapi Daily Activity Report Generator
Fetches data from Google Sheets and generates a self-contained HTML dashboard.
"""

import csv
import sys
import io
import json
import argparse
import requests
import tempfile
import os
from datetime import datetime, timedelta
from urllib.parse import urlencode
from typing import Dict, List, Tuple, Optional, Any


class GoogleSheetsFetcher:
    """Handles fetching and parsing Google Sheets CSV data."""

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def get_csv_url(self, sheet_id: str, sheet_name: str, cell_range: str = None) -> str:
        """Generate CSV export URL for a Google Sheet, optionally with a cell range."""
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&sheet={sheet_name}"
        if cell_range:
            url += f"&range={cell_range}"
        return url

    def fetch_sheet(self, sheet_id: str, sheet_name: str, cell_range: str = None,
                    gid: str = None) -> Optional[List[List[str]]]:
        """Fetch and parse a Google Sheet as CSV data.

        Args:
            cell_range: Optional cell range (e.g., 'J19:O28') for targeted fetching.
                        Needed for areas with merged cells that don't export correctly
                        in full-sheet CSV mode.
            gid: Optional sheet GID.  When provided the raw export?format=csv
                 endpoint is used instead of the gviz/tq endpoint.
        """
        try:
            if gid is not None:
                url = (f"https://docs.google.com/spreadsheets/d/{sheet_id}"
                       f"/export?format=csv&gid={gid}")
            else:
                url = self.get_csv_url(sheet_id, sheet_name, cell_range)
            response = self.session.get(url, timeout=30)
            response.encoding = 'utf-8'
            response.raise_for_status()

            reader = csv.reader(io.StringIO(response.text))
            data = list(reader)
            return data
        except Exception as e:
            range_label = f" range '{cell_range}'" if cell_range else ""
            print(f"Error fetching sheet '{sheet_name}'{range_label}: {e}", file=sys.stderr)
            return None

    def fetch_sheet_xlsx(self, sheet_id: str) -> Optional[Any]:
        """Fetch a Google Sheet as XLSX to access cell formatting (e.g., background colors)."""
        try:
            url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
            response = self.session.get(url, timeout=60)
            response.raise_for_status()

            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp.write(response.content)
            tmp.close()

            from openpyxl import load_workbook
            wb = load_workbook(tmp.name, data_only=True)
            os.unlink(tmp.name)
            return wb
        except Exception as e:
            print(f"Error fetching XLSX: {e}", file=sys.stderr)
            return None

    def get_current_week_sheet_name(self) -> str:
        """Get the current calendar week sheet name (e.g., 'CW10 2026')."""
        today = datetime.now()
        iso_calendar = today.isocalendar()
        week_number = iso_calendar[1]
        year = iso_calendar[0]
        return f"CW{week_number} {year}"

    def resolve_sheet_name_from_wb(self, wb, expected_name: str) -> str:
        """Find exact sheet name in XLSX workbook, handling whitespace mismatches."""
        if not wb:
            return expected_name
        for name in wb.sheetnames:
            if name.strip() == expected_name.strip():
                if name != expected_name:
                    print(f"  Resolved sheet name: '{expected_name}' -> '{name}' (whitespace adjusted)")
                return name
        print(f"  Warning: Sheet '{expected_name}' not found in workbook")
        return expected_name


class SalesHuddleParser:
    """Parses Sales Huddle sheet data."""

    def __init__(self, sheet_data: List[List[str]], hot_deals_ranges: Dict[str, Optional[List[List[str]]]] = None):
        self.data = sheet_data
        # Hot deals range data fetched separately to work around merged cell CSV export issues.
        # Keys: 'outbound', 'inbound', 'intl_inbound', 'intl_outbound'
        self.hot_deals_ranges = hot_deals_ranges or {}

    def get_cell(self, row: int, col: int) -> str:
        """Safely get a cell value, handling indices."""
        try:
            if row < len(self.data) and col < len(self.data[row]):
                return self.data[row][col].strip()
            return ""
        except:
            return ""

    def _parse_hot_deals_from_range(self, range_data, agent1_col, agent2_col, agent1_name, agent2_name,
                                     agent3_col=None, agent3_name=None):
        """Parse hot deals from a range-specific CSV fetch.

        Range data has columns relative to the fetched range (0-indexed).
        agent1_col/agent2_col are column indices within the range data.
        Optionally supports a third agent (agent3_col/agent3_name).
        """
        hot_deals = {
            "hot_deal": {agent1_name: [], agent2_name: []},
            "ctp": {agent1_name: [], agent2_name: []},
            "won": {agent1_name: [], agent2_name: []}
        }
        if agent3_name:
            for cat in hot_deals:
                hot_deals[cat][agent3_name] = []

        if not range_data:
            return hot_deals

        current_category = None
        category_keywords = {"hot deal", "ctp", "won"}
        for row in range_data:
            agent1_val = row[agent1_col].strip() if agent1_col < len(row) else ""
            agent2_val = row[agent2_col].strip() if agent2_col < len(row) else ""
            agent3_val = ""
            if agent3_col is not None and agent3_name:
                agent3_val = row[agent3_col].strip() if agent3_col < len(row) else ""

            check1 = agent1_val.lower() if agent1_val else ""
            check2 = agent2_val.lower() if agent2_val else ""
            check3 = agent3_val.lower() if agent3_val else ""

            if check1 == "hot deal" or check2 == "hot deal" or check3 == "hot deal":
                current_category = "hot_deal"
                # Pick up non-category values on the same row
                for val, chk, name in [(agent1_val, check1, agent1_name),
                                        (agent2_val, check2, agent2_name),
                                        (agent3_val, check3, agent3_name if agent3_name else None)]:
                    if name and val and chk not in category_keywords:
                        hot_deals[current_category][name].append(val)
                continue
            elif check1 == "ctp" or check2 == "ctp" or check3 == "ctp":
                current_category = "ctp"
                for val, chk, name in [(agent1_val, check1, agent1_name),
                                        (agent2_val, check2, agent2_name),
                                        (agent3_val, check3, agent3_name if agent3_name else None)]:
                    if name and val and chk not in category_keywords:
                        hot_deals[current_category][name].append(val)
                continue
            elif check1 == "won" or check2 == "won" or check3 == "won":
                current_category = "won"
                for val, chk, name in [(agent1_val, check1, agent1_name),
                                        (agent2_val, check2, agent2_name),
                                        (agent3_val, check3, agent3_name if agent3_name else None)]:
                    if name and val and chk not in category_keywords:
                        hot_deals[current_category][name].append(val)
                continue

            if current_category:
                if agent1_val and agent1_val.lower() not in category_keywords:
                    hot_deals[current_category][agent1_name].append(agent1_val)
                if agent2_val and agent2_val.lower() not in category_keywords:
                    hot_deals[current_category][agent2_name].append(agent2_val)
                if agent3_name and agent3_val and agent3_val.lower() not in category_keywords:
                    hot_deals[current_category][agent3_name].append(agent3_val)

        return hot_deals

    def get_date_info(self) -> Dict[str, str]:
        """Extract date information from header section."""
        try:
            date_str = self.get_cell(1, 2)  # Row 2, Col C
            return {"date": date_str}
        except:
            return {"date": datetime.now().strftime("%d/%m/%Y")}

    def parse_outbound_section(self) -> Dict[str, Any]:
        """Parse the Outbound sales section."""
        result = {
            "agents": {},
            "funnel": [],
            "raw_data": []
        }

        try:
            # Outbound section starts at row 9 (index 9)
            # Agents: Yayee (J-L), Toey (M-O)
            # Row 12: Activities (Calls)
            # Row 13: Contact
            # Row 14: Demo Scheduled
            # Row 15: Demo Attended
            # Row 16: Contact to Demo %
            # Row 17: Won

            funnel_metrics = ["Activities (Calls)", "Contact", "Demo Scheduled", "Demo Attended", "Won"]
            funnel_rows = [12, 13, 14, 15, 17]  # Row indices for each metric

            # Extract target and total WTD
            target_wtd = self.get_cell(10, 2)  # Row 11, Col C (Target WTD Total)
            total_wtd = self.get_cell(11, 6)   # Row 12, Col G (Total WTD)
            total_daily = self.get_cell(11, 7)  # Row 12, Col H (Total Daily)
            total_vs_target = self.get_cell(11, 8)  # Row 12, Col I (WTD vs Target)

            result["summary"] = {
                "target_wtd": target_wtd,
                "total_wtd": total_wtd,
                "total_daily": total_daily,
                "total_vs_target": total_vs_target
            }

            # Parse each metric
            for metric_name, row_idx in zip(funnel_metrics, funnel_rows):
                metric_data = {
                    "name": metric_name,
                    "total_wtd": self.get_cell(row_idx, 6),
                    "total_daily": self.get_cell(row_idx, 7),
                    "total_vs_target": self.get_cell(row_idx, 8),
                    "yayee_wtd": self.get_cell(row_idx, 9),
                    "yayee_daily": self.get_cell(row_idx, 10),
                    "yayee_vs_target": self.get_cell(row_idx, 11),
                    "toey_wtd": self.get_cell(row_idx, 12),
                    "toey_daily": self.get_cell(row_idx, 13),
                    "toey_vs_target": self.get_cell(row_idx, 14)
                }
                result["funnel"].append(metric_data)

            # Parse Hot Deals using range-specific data (merged cell CSV workaround)
            # Range fetch: J19:O28 -> col 0 = Yayee (J), col 3 = Toey (M)
            range_data = self.hot_deals_ranges.get("outbound")
            if range_data:
                result["hot_deals"] = self._parse_hot_deals_from_range(
                    range_data, agent1_col=0, agent2_col=3,
                    agent1_name="yayee", agent2_name="toey"
                )
            else:
                hot_deals = {
                    "hot_deal": {"yayee": [], "toey": []},
                    "ctp": {"yayee": [], "toey": []},
                    "won": {"yayee": [], "toey": []}
                }
                current_category = None
                for row_idx in range(18, min(len(self.data), 30)):
                    yayee_val = self.get_cell(row_idx, 9)
                    toey_val = self.get_cell(row_idx, 12)
                    check_val = yayee_val.lower().strip() if yayee_val else ""
                    check_val2 = toey_val.lower().strip() if toey_val else ""
                    if check_val == "hot deal" or check_val2 == "hot deal":
                        current_category = "hot_deal"
                        continue
                    elif check_val == "ctp" or check_val2 == "ctp":
                        current_category = "ctp"
                        continue
                    elif check_val == "won" or check_val2 == "won":
                        current_category = "won"
                        continue
                    if current_category:
                        if yayee_val and yayee_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["yayee"].append(yayee_val)
                        if toey_val and toey_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["toey"].append(toey_val)
                result["hot_deals"] = hot_deals

        except Exception as e:
            print(f"Error parsing outbound section: {e}", file=sys.stderr)

        return result

    def parse_inbound_section(self) -> Dict[str, Any]:
        """Parse the Inbound sales section."""
        result = {
            "funnel": [],
            "hot_deals": {},
            "summary": {}
        }

        try:
            # Inbound section around row 9, columns R onwards
            # Agents: Pleum (AA-AC), Loogpad (AD-AF)
            # Same funnel rows as outbound

            funnel_metrics = ["Activities (Calls)", "Contact", "Demo Scheduled", "Demo Attended", "Won"]
            funnel_rows = [12, 13, 14, 15, 17]

            # Column offsets for Inbound
            total_wtd_col = 23     # X
            total_daily_col = 24   # Y
            total_vs_target_col = 25  # Z
            pleum_wtd_col = 26     # AA
            pleum_daily_col = 27   # AB
            pleum_vs_target_col = 28  # AC
            loogpad_wtd_col = 29   # AD
            loogpad_daily_col = 30 # AE
            loogpad_vs_target_col = 31  # AF
            pear_wtd_col = 32      # AG
            pear_daily_col = 33    # AH
            pear_vs_target_col = 34  # AI

            # Extract summary
            target_wtd = self.get_cell(10, 19)  # T
            result["summary"] = {
                "target_wtd": target_wtd,
                "total_wtd": self.get_cell(11, total_wtd_col),
                "total_daily": self.get_cell(11, total_daily_col),
                "total_vs_target": self.get_cell(11, total_vs_target_col)
            }

            # Parse funnel metrics
            for metric_name, row_idx in zip(funnel_metrics, funnel_rows):
                metric_data = {
                    "name": metric_name,
                    "total_wtd": self.get_cell(row_idx, total_wtd_col),
                    "total_daily": self.get_cell(row_idx, total_daily_col),
                    "total_vs_target": self.get_cell(row_idx, total_vs_target_col),
                    "pleum_wtd": self.get_cell(row_idx, pleum_wtd_col),
                    "pleum_daily": self.get_cell(row_idx, pleum_daily_col),
                    "pleum_vs_target": self.get_cell(row_idx, pleum_vs_target_col),
                    "pear_wtd": self.get_cell(row_idx, pear_wtd_col),
                    "pear_daily": self.get_cell(row_idx, pear_daily_col),
                    "pear_vs_target": self.get_cell(row_idx, pear_vs_target_col),
                    "loogpad_wtd": self.get_cell(row_idx, loogpad_wtd_col),
                    "loogpad_daily": self.get_cell(row_idx, loogpad_daily_col),
                    "loogpad_vs_target": self.get_cell(row_idx, loogpad_vs_target_col)
                }
                result["funnel"].append(metric_data)

            # Parse Hot Deals using range-specific data (merged cell CSV workaround)
            # Range fetch: AA19:AI55 -> col 0 = Pleum (AA), col 3 = Loogpad (AD), col 6 = Pear (AG)
            range_data = self.hot_deals_ranges.get("inbound")
            if range_data:
                result["hot_deals"] = self._parse_hot_deals_from_range(
                    range_data, agent1_col=0, agent2_col=3,
                    agent1_name="pleum", agent2_name="loogpad",
                    agent3_col=6, agent3_name="pear"
                )
            else:
                hot_deals = {
                    "hot_deal": {"pleum": [], "loogpad": [], "pear": []},
                    "ctp": {"pleum": [], "loogpad": [], "pear": []},
                    "won": {"pleum": [], "loogpad": [], "pear": []}
                }
                current_category = None
                for row_idx in range(18, min(len(self.data), 30)):
                    pleum_val = self.get_cell(row_idx, pleum_wtd_col)
                    loogpad_val = self.get_cell(row_idx, loogpad_wtd_col)
                    pear_val = self.get_cell(row_idx, pear_wtd_col)
                    check_val = pleum_val.lower().strip() if pleum_val else ""
                    check_val2 = loogpad_val.lower().strip() if loogpad_val else ""
                    check_val3 = pear_val.lower().strip() if pear_val else ""
                    if check_val == "hot deal" or check_val2 == "hot deal" or check_val3 == "hot deal":
                        current_category = "hot_deal"
                        continue
                    elif check_val == "ctp" or check_val2 == "ctp" or check_val3 == "ctp":
                        current_category = "ctp"
                        continue
                    elif check_val == "won" or check_val2 == "won" or check_val3 == "won":
                        current_category = "won"
                        continue
                    if current_category:
                        if pleum_val and pleum_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["pleum"].append(pleum_val)
                        if loogpad_val and loogpad_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["loogpad"].append(loogpad_val)
                        if pear_val and pear_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["pear"].append(pear_val)
                result["hot_deals"] = hot_deals

        except Exception as e:
            print(f"Error parsing inbound section: {e}", file=sys.stderr)

        return result

    def parse_renewal_section(self, renewal_range_data=None, green_cells=None) -> Dict[str, Any]:
        """Parse the Account Management - Renewal section.

        Args:
            renewal_range_data: CSV data from range R58:AF100 (optional, uses range-specific fetch)
            green_cells: Set of account names that have green background (confirmed to renew)
        """
        result = {
            "header_date": "",
            "won": {},
            "due_to_renew": {"pleum": [], "loogpad": []},
            "renewed": {"pleum": [], "loogpad": []},
            "churned": {"pleum": [], "loogpad": []}
        }

        if green_cells is None:
            green_cells = set()

        data = renewal_range_data if renewal_range_data else self.data

        try:
            if not data or len(data) == 0:
                return result

            # If using range data (R58:AF), columns are 0-indexed within range
            # Col 0=R, 1=S, ..., 5=W, 6=X, 7=Y, 9=AA, 10=AB, 12=AD, 13=AE

            def get_val(row_idx, col_idx):
                if row_idx < len(data) and col_idx < len(data[row_idx]):
                    return data[row_idx][col_idx].strip()
                return ""

            # Row 0: Header - extract date
            header = get_val(0, 0)
            if header:
                # Extract date from "Account Management - Renewal 21/04/2026"
                parts = header.split()
                for part in parts:
                    if "/" in part and len(part) >= 8:
                        result["header_date"] = part
                        break

            # Row 1: Won metrics
            won_label = get_val(1, 5)
            if won_label.lower().strip() == "won":
                result["won"] = {
                    "total_wtd": get_val(1, 6),
                    "total_daily": get_val(1, 7),
                    "pleum_wtd": get_val(1, 9),
                    "pleum_daily": get_val(1, 10),
                    "loogpad_wtd": get_val(1, 12),
                    "loogpad_daily": get_val(1, 13)
                }

            # Parse sections: due_to_renew, renewed, churned
            current_section = "due_to_renew"  # Default starting section

            for row_idx in range(2, len(data)):
                # Check for section headers
                label_col1 = get_val(row_idx, 1).lower().strip()

                if "due to renew" in label_col1:
                    current_section = "due_to_renew"
                    continue
                elif "renewed" in label_col1 and "not" not in label_col1:
                    current_section = "renewed"
                    continue
                elif "churned" in label_col1:
                    current_section = "churned"
                    continue
                elif "have not reached" in label_col1 or "not reached" in label_col1:
                    break  # Stop parsing after this

                if current_section == "due_to_renew":
                    # Pleum: date in col9, name in col10
                    pleum_date = get_val(row_idx, 9)
                    pleum_name = get_val(row_idx, 10)
                    if pleum_name:
                        confirmed = pleum_name in green_cells
                        result["due_to_renew"]["pleum"].append({
                            "name": pleum_name,
                            "date": pleum_date,
                            "confirmed": confirmed
                        })

                    # Loogpad: date in col12, name in col13
                    loogpad_date = get_val(row_idx, 12)
                    loogpad_name = get_val(row_idx, 13)
                    if loogpad_name:
                        confirmed = loogpad_name in green_cells
                        result["due_to_renew"]["loogpad"].append({
                            "name": loogpad_name,
                            "date": loogpad_date,
                            "confirmed": confirmed
                        })

                elif current_section == "renewed":
                    # Renewed: name in col9 (Pleum), col12 (Loogpad) - no dates
                    pleum_name = get_val(row_idx, 9)
                    loogpad_name = get_val(row_idx, 12)
                    if pleum_name:
                        result["renewed"]["pleum"].append(pleum_name)
                    if loogpad_name:
                        result["renewed"]["loogpad"].append(loogpad_name)

                elif current_section == "churned":
                    pleum_name = get_val(row_idx, 9)
                    loogpad_name = get_val(row_idx, 12)
                    if pleum_name:
                        result["churned"]["pleum"].append(pleum_name)
                    if loogpad_name:
                        result["churned"]["loogpad"].append(loogpad_name)

        except Exception as e:
            print(f"Error parsing renewal section: {e}", file=sys.stderr)

        return result

    def parse_intl_inbound_section(self) -> Dict[str, Any]:
        """Parse the International Inbound sales section."""
        result = {
            "funnel": [],
            "hot_deals": {},
            "summary": {}
        }

        try:
            # International Inbound section at row 10
            # Agents: Sheronika, Thanom
            # Funnel metrics at rows 13-18 (6 metrics)

            funnel_rows = [13, 14, 15, 16, 17, 18]

            # Column offsets for International Inbound
            label_col = 38         # AM - metric labels
            total_wtd_col = 43     # AR
            total_daily_col = 44   # AS
            total_vs_target_col = 45  # AT
            sheronika_wtd_col = 46    # AU
            sheronika_daily_col = 47  # AV
            sheronika_vs_target_col = 48  # AW
            thanom_wtd_col = 49    # AX
            thanom_daily_col = 50  # AY
            thanom_vs_target_col = 51  # AZ

            # Extract summary
            target_wtd = self.get_cell(10, 38)  # Row 11, col AM
            result["summary"] = {
                "target_wtd": target_wtd,
                "total_wtd": self.get_cell(11, total_wtd_col),
                "total_daily": self.get_cell(11, total_daily_col),
                "total_vs_target": self.get_cell(11, total_vs_target_col)
            }

            # Parse funnel metrics - read metric names dynamically from label column
            for row_idx in funnel_rows:
                metric_name = self.get_cell(row_idx, label_col)
                if not metric_name:
                    metric_name = f"Metric Row {row_idx}"

                metric_data = {
                    "name": metric_name,
                    "total_wtd": self.get_cell(row_idx, total_wtd_col),
                    "total_daily": self.get_cell(row_idx, total_daily_col),
                    "total_vs_target": self.get_cell(row_idx, total_vs_target_col),
                    "sheronika_wtd": self.get_cell(row_idx, sheronika_wtd_col),
                    "sheronika_daily": self.get_cell(row_idx, sheronika_daily_col),
                    "sheronika_vs_target": self.get_cell(row_idx, sheronika_vs_target_col),
                    "thanom_wtd": self.get_cell(row_idx, thanom_wtd_col),
                    "thanom_daily": self.get_cell(row_idx, thanom_daily_col),
                    "thanom_vs_target": self.get_cell(row_idx, thanom_vs_target_col)
                }
                result["funnel"].append(metric_data)

            # Parse Hot Deals using range-specific data (merged cell CSV workaround)
            # Range fetch: AU19:AZ28 -> col 0 = Sheronika (AU), col 3 = Thanom (AX)
            range_data = self.hot_deals_ranges.get("intl_inbound")
            if range_data:
                result["hot_deals"] = self._parse_hot_deals_from_range(
                    range_data, agent1_col=0, agent2_col=3,
                    agent1_name="sheronika", agent2_name="thanom"
                )
            else:
                hot_deals = {
                    "hot_deal": {"sheronika": [], "thanom": []},
                    "ctp": {"sheronika": [], "thanom": []},
                    "won": {"sheronika": [], "thanom": []}
                }
                current_category = None
                for row_idx in range(18, min(len(self.data), 30)):
                    sheronika_val = self.get_cell(row_idx, sheronika_wtd_col)
                    thanom_val = self.get_cell(row_idx, thanom_wtd_col)
                    check_val = sheronika_val.lower().strip() if sheronika_val else ""
                    check_val2 = thanom_val.lower().strip() if thanom_val else ""
                    if check_val == "hot deal" or check_val2 == "hot deal":
                        current_category = "hot_deal"
                        continue
                    elif check_val == "ctp" or check_val2 == "ctp":
                        current_category = "ctp"
                        continue
                    elif check_val == "won" or check_val2 == "won":
                        current_category = "won"
                        continue
                    if current_category:
                        if sheronika_val and sheronika_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["sheronika"].append(sheronika_val)
                        if thanom_val and thanom_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["thanom"].append(thanom_val)
                result["hot_deals"] = hot_deals

        except Exception as e:
            print(f"Error parsing international inbound section: {e}", file=sys.stderr)

        return result

    def parse_intl_outbound_section(self) -> Dict[str, Any]:
        """Parse the International Outbound sales section."""
        result = {
            "funnel": [],
            "summary": {},
            "hot_deals": {}
        }

        try:
            # International Outbound section
            # Agents: Sheronika, Thanom
            # Funnel metrics at rows 13-18 (6 metrics)

            funnel_rows = [13, 14, 15, 16, 17, 18]

            # Column offsets for International Outbound
            label_col = 54         # BC - metric labels
            total_wtd_col = 59     # BH
            total_daily_col = 60   # BI
            total_vs_target_col = 61  # BJ
            sheronika_wtd_col = 62    # BK
            sheronika_daily_col = 63  # BL
            sheronika_vs_target_col = 64  # BM
            thanom_wtd_col = 65    # BN
            thanom_daily_col = 66  # BO
            thanom_vs_target_col = 67  # BP

            # Extract summary
            target_wtd = self.get_cell(10, 54)  # Row 11, col BC
            result["summary"] = {
                "target_wtd": target_wtd,
                "total_wtd": self.get_cell(11, total_wtd_col),
                "total_daily": self.get_cell(11, total_daily_col),
                "total_vs_target": self.get_cell(11, total_vs_target_col)
            }

            # Parse funnel metrics - read metric names dynamically from label column
            for row_idx in funnel_rows:
                metric_name = self.get_cell(row_idx, label_col)
                if not metric_name:
                    metric_name = f"Metric Row {row_idx}"

                metric_data = {
                    "name": metric_name,
                    "total_wtd": self.get_cell(row_idx, total_wtd_col),
                    "total_daily": self.get_cell(row_idx, total_daily_col),
                    "total_vs_target": self.get_cell(row_idx, total_vs_target_col),
                    "sheronika_wtd": self.get_cell(row_idx, sheronika_wtd_col),
                    "sheronika_daily": self.get_cell(row_idx, sheronika_daily_col),
                    "sheronika_vs_target": self.get_cell(row_idx, sheronika_vs_target_col),
                    "thanom_wtd": self.get_cell(row_idx, thanom_wtd_col),
                    "thanom_daily": self.get_cell(row_idx, thanom_daily_col),
                    "thanom_vs_target": self.get_cell(row_idx, thanom_vs_target_col)
                }
                result["funnel"].append(metric_data)

            # Parse Hot Deals using range-specific data (merged cell CSV workaround)
            # Range fetch: BK19:BP28 -> col 0 = Sheronika (BK), col 3 = Thanom (BN)
            range_data = self.hot_deals_ranges.get("intl_outbound")
            if range_data:
                result["hot_deals"] = self._parse_hot_deals_from_range(
                    range_data, agent1_col=0, agent2_col=3,
                    agent1_name="sheronika", agent2_name="thanom"
                )
            else:
                hot_deals = {
                    "hot_deal": {"sheronika": [], "thanom": []},
                    "ctp": {"sheronika": [], "thanom": []},
                    "won": {"sheronika": [], "thanom": []}
                }
                current_category = None
                for row_idx in range(18, min(len(self.data), 30)):
                    sheronika_val = self.get_cell(row_idx, sheronika_wtd_col)
                    thanom_val = self.get_cell(row_idx, thanom_wtd_col)
                    check_val = sheronika_val.lower().strip() if sheronika_val else ""
                    check_val2 = thanom_val.lower().strip() if thanom_val else ""
                    if check_val == "hot deal" or check_val2 == "hot deal":
                        current_category = "hot_deal"
                        continue
                    elif check_val == "ctp" or check_val2 == "ctp":
                        current_category = "ctp"
                        continue
                    elif check_val == "won" or check_val2 == "won":
                        current_category = "won"
                        continue
                    if current_category:
                        if sheronika_val and sheronika_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["sheronika"].append(sheronika_val)
                        if thanom_val and thanom_val.lower() not in ["hot deal", "ctp", "won"]:
                            hot_deals[current_category]["thanom"].append(thanom_val)
                result["hot_deals"] = hot_deals

        except Exception as e:
            print(f"Error parsing international outbound section: {e}", file=sys.stderr)

        return result


class RegistrationWeeklyParser:
    """Parses the registration_weekly tab of the Ads Data sheet.

    Source sheet: 1s5AC58mAylpSDknU7L7HRJUPrVf36b0TvzD35tW-Wdw
    Tab: registration_weekly

    Columns (1-indexed):
      A cw, B week_start_mon, C week_end_sun, D region, E market,
      F ad_source, G ad_campaign_id,
      H verified, I integrated, J qualified, K highly_qualified,
      L premium, M best, N hqplus, O total

    Aggregates the current ISO-week rows into:
      - GLOBAL (grand total)
      - TH, SEA, ROW (region rollups)
      - MY, SG, PH (SEA submarkets)
      - ROW submarkets by country (BR, DE, CH, ‚Äö√Ñ¬∂)

    NOTE: We bucket by *market* (column E), not the region column, because the
    current-week feed sometimes tags rows as region=ROW with country names like
    THAILAND/MALAYSIA in the market column. Bucketing by market makes the
    dashboard self-correcting against that tagging issue.

    For each bucket we compute:
      - qualified           = sum(col J)
      - hqplus              = sum(col N)
      - total               = qualified + hqplus   (per spec ‚Äö√Ñ√Æ not col O)
      - attributed          = total contribution from rows with ad_source set
      - unattributed        = total contribution from rows with blank/unknown
                              ad_source
    Attributed + Unattributed = Total.
    """

    # Markets that belong to the TH or SEA buckets. Anything else falls into ROW.
    # Accepts both 2-letter codes and full country names (uppercased).
    TH_MARKETS = {"TH", "THAILAND"}
    SEA_SUBMARKET_MAP = {
        "MY": "MY", "MALAYSIA": "MY",
        "SG": "SG", "SINGAPORE": "SG",
        "PH": "PH", "PHILIPPINES": "PH",
    }
    SEA_OTHER_MARKETS = {
        "ID", "INDONESIA",
        "VN", "VIETNAM",
        "MM", "MYANMAR",
        "KH", "CAMBODIA",
        "LA", "LAOS",
        "BN", "BRUNEI",
        "TL", "TP", "EAST TIMOR",
    }

    # Country name ‚Äö√ú√≠ ISO alpha-2 code for ROW (and a few SEA aliases). Used to
    # label ROW sub-rows. If a market value is already a 2-letter code we use
    # it directly.
    COUNTRY_NAME_TO_CODE = {
        # SEA aliases (defensive)
        "THAILAND": "TH", "MALAYSIA": "MY", "SINGAPORE": "SG",
        "PHILIPPINES": "PH", "INDONESIA": "ID", "VIETNAM": "VN",
        "MYANMAR": "MM", "CAMBODIA": "KH", "LAOS": "LA", "BRUNEI": "BN",
        "EAST TIMOR": "TL",
        # ROW
        "INDIA": "IN",
        "UNITED STATES OF AMERICA": "US", "UNITED STATES": "US", "USA": "US",
        "UNITED KINGDOM": "UK", "GREAT BRITAIN": "UK",
        "GERMANY": "DE", "SWITZERLAND": "CH", "AUSTRIA": "AT",
        "FRANCE": "FR", "NETHERLANDS": "NL", "BELGIUM": "BE",
        "ITALY": "IT", "SPAIN": "ES", "PORTUGAL": "PT", "POLAND": "PL",
        "SWEDEN": "SE", "NORWAY": "NO", "DENMARK": "DK", "FINLAND": "FI",
        "IRELAND": "IE", "GREECE": "GR", "CZECH REPUBLIC": "CZ",
        "HUNGARY": "HU", "ROMANIA": "RO", "BULGARIA": "BG",
        "UKRAINE": "UA", "RUSSIA": "RU", "TURKEY": "TR",
        "JAPAN": "JP", "CHINA": "CN", "SOUTH KOREA": "KR", "KOREA": "KR",
        "HONG KONG": "HK", "TAIWAN": "TW", "MACAU": "MO", "MONGOLIA": "MN",
        "BANGLADESH": "BD", "PAKISTAN": "PK", "SRI LANKA": "LK",
        "NEPAL": "NP", "AFGHANISTAN": "AF",
        "KAZAKHSTAN": "KZ", "KYRGYZSTAN": "KG", "UZBEKISTAN": "UZ",
        "AUSTRALIA": "AU", "NEW ZEALAND": "NZ",
        "CANADA": "CA", "MEXICO": "MX",
        "BRAZIL": "BR", "ARGENTINA": "AR", "CHILE": "CL",
        "COLOMBIA": "CO", "PERU": "PE", "VENEZUELA": "VE",
        "ECUADOR": "EC", "URUGUAY": "UY",
        "SOUTH AFRICA": "ZA", "NIGERIA": "NG", "EGYPT": "EG",
        "KENYA": "KE", "MOROCCO": "MA", "GHANA": "GH",
        "ALGERIA": "DZ", "TUNISIA": "TN", "ETHIOPIA": "ET",
        "UGANDA": "UG", "TANZANIA": "TZ",
        "UNITED ARAB EMIRATES": "AE", "UAE": "AE",
        "SAUDI ARABIA": "SA", "ISRAEL": "IL",
        "IRAN": "IR", "IRAQ": "IQ", "JORDAN": "JO", "LEBANON": "LB",
        "SYRIA": "SY", "YEMEN": "YE",
        "QATAR": "QA", "KUWAIT": "KW", "BAHRAIN": "BH", "OMAN": "OM",
        "PALAU": "PW", "FIJI": "FJ",
        "UNKNOWN": "??",
    }

    # Reverse map (code ‚Äö√ú√≠ friendly label) for display.
    CODE_TO_LABEL = {
        "US": "United States", "UK": "United Kingdom",
        "DE": "Germany", "CH": "Switzerland",
        "FR": "France", "NL": "Netherlands", "BE": "Belgium",
        "IT": "Italy", "ES": "Spain", "PT": "Portugal", "PL": "Poland",
        "SE": "Sweden", "NO": "Norway", "DK": "Denmark", "FI": "Finland",
        "AT": "Austria",
        "IN": "India", "PK": "Pakistan", "BD": "Bangladesh",
        "NP": "Nepal", "LK": "Sri Lanka", "AF": "Afghanistan",
        "JP": "Japan", "CN": "China", "KR": "South Korea",
        "HK": "Hong Kong", "TW": "Taiwan", "MO": "Macau", "MN": "Mongolia",
        "KZ": "Kazakhstan", "KG": "Kyrgyzstan", "UZ": "Uzbekistan",
        "AU": "Australia", "NZ": "New Zealand",
        "CA": "Canada", "MX": "Mexico",
        "BR": "Brazil", "AR": "Argentina", "CL": "Chile",
        "CO": "Colombia", "PE": "Peru", "VE": "Venezuela",
        "EC": "Ecuador", "UY": "Uruguay",
        "ZA": "South Africa", "NG": "Nigeria", "EG": "Egypt",
        "KE": "Kenya", "MA": "Morocco", "GH": "Ghana",
        "DZ": "Algeria", "TN": "Tunisia", "ET": "Ethiopia",
        "UG": "Uganda", "TZ": "Tanzania",
        "AE": "UAE", "SA": "Saudi Arabia", "IL": "Israel",
        "IR": "Iran", "IQ": "Iraq", "JO": "Jordan", "LB": "Lebanon",
        "SY": "Syria", "YE": "Yemen",
        "QA": "Qatar", "KW": "Kuwait", "BH": "Bahrain", "OM": "Oman",
        "PW": "Palau", "FJ": "Fiji",
        "TR": "Turkey", "RU": "Russia", "UA": "Ukraine",
        "GR": "Greece", "CZ": "Czechia", "HU": "Hungary",
        "RO": "Romania", "BG": "Bulgaria", "IE": "Ireland",
        "??": "Unknown",
    }

    # ad_source values that mean "we don't actually know the source".
    UNATTRIBUTED_AD_SOURCES = {"", "UNKNOWN", "NONE", "N/A"}

    def __init__(self, sheet_data: List[List[str]], week_start_mon: str):
        self.data = sheet_data
        self.week_start_mon = week_start_mon

    @staticmethod
    def _to_int(value: str) -> int:
        try:
            return int(float(str(value).replace(",", "").strip() or "0"))
        except (ValueError, TypeError):
            return 0

    # Bucket logic ---------------------------------------------------------

    def _classify_market(self, market: str):
        """Return (region, submarket, country_code).

        region        ‚Äö√†√† {"TH", "SEA", "ROW"}
        submarket     ‚Äö√†√† {"MY", "SG", "PH"} or None  (only set for SEA majors)
        country_code  ISO alpha-2 (or fallback) ‚Äö√Ñ√Æ only meaningful for ROW;
                      for SEA/TH this is left to the caller (we don't break
                      out non-major SEA countries).
        """
        if not market:
            return ("ROW", None, "??")
        key = market.strip().upper()
        if key in self.TH_MARKETS:
            return ("TH", None, "TH")
        if key in self.SEA_SUBMARKET_MAP:
            sub = self.SEA_SUBMARKET_MAP[key]
            return ("SEA", sub, sub)
        if key in self.SEA_OTHER_MARKETS:
            return ("SEA", None, key if len(key) == 2 else
                    self.COUNTRY_NAME_TO_CODE.get(key, key[:2]))
        # ROW
        if len(key) == 2 and key.isalpha():
            code = key
        else:
            code = self.COUNTRY_NAME_TO_CODE.get(key)
            if not code:
                # Fallback: first 2 alpha chars uppercase, or "??".
                code = "".join(c for c in key if c.isalpha())[:2].upper() or "??"
        return ("ROW", None, code)

    def _label_for_country(self, code: str, raw_market: str) -> str:
        label = self.CODE_TO_LABEL.get(code)
        if label:
            return label
        # Fallback to titlecasing the raw market name.
        return raw_market.strip().title() if raw_market else code

    def _is_attributed(self, ad_source: str) -> bool:
        return ad_source.strip().upper() not in self.UNATTRIBUTED_AD_SOURCES

    # Aggregation ---------------------------------------------------------

    @staticmethod
    def _empty_bucket():
        return {
            "qualified_wtd": 0,
            "hqplus_wtd": 0,
            "total_wtd": 0,          # = qualified + hqplus per spec
            "attributed_wtd": 0,
            "unattributed_wtd": 0,
        }

    @classmethod
    def _stringify(cls, bucket: Dict[str, int]) -> Dict[str, str]:
        return {k: str(v) for k, v in bucket.items()}

    def parse_data(self) -> Dict[str, Any]:
        # Fixed-order buckets always present in the table.
        fixed_regions = ["GLOBAL", "TH", "SEA", "MY", "SG", "PH", "ROW"]
        regions: Dict[str, Dict[str, int]] = {
            r: self._empty_bucket() for r in fixed_regions
        }
        # ROW sub-buckets discovered dynamically by country code.
        row_countries: Dict[str, Dict[str, Any]] = {}
        # Track best raw_market label seen for each ROW country code.
        row_country_labels: Dict[str, str] = {}

        result = {
            "date": self.week_start_mon,
            "regions": regions,
            "row_countries": [],   # ordered list of {code,label,...stats}
            "total": self._empty_bucket(),
        }

        if not self.data or len(self.data) < 2:
            print("Warning: registration_weekly is empty/header-only",
                  file=sys.stderr)
            return self._finalize(result, row_countries, row_country_labels)

        header = [c.strip().lower() for c in self.data[0]]
        try:
            idx_week = header.index("week_start_mon")
            idx_market = header.index("market")
            idx_adsrc = header.index("ad_source")
            idx_qualified = header.index("qualified")
            idx_hqplus = header.index("hqplus")
        except ValueError as e:
            print(f"Error: registration_weekly missing column: {e}",
                  file=sys.stderr)
            return self._finalize(result, row_countries, row_country_labels)

        max_idx = max(idx_week, idx_market, idx_adsrc,
                      idx_qualified, idx_hqplus)
        matched = 0
        for row in self.data[1:]:
            if len(row) <= max_idx:
                continue
            if row[idx_week].strip() != self.week_start_mon:
                continue
            matched += 1
            market = row[idx_market]
            ad_source = row[idx_adsrc]
            q = self._to_int(row[idx_qualified])
            h = self._to_int(row[idx_hqplus])
            t = q + h                       # Total per spec
            attributed = self._is_attributed(ad_source)
            a = t if attributed else 0
            u = 0 if attributed else t

            region, submarket, code = self._classify_market(market)

            def _add(bucket):
                bucket["qualified_wtd"] += q
                bucket["hqplus_wtd"] += h
                bucket["total_wtd"] += t
                bucket["attributed_wtd"] += a
                bucket["unattributed_wtd"] += u

            _add(regions["GLOBAL"])
            _add(regions[region])
            if submarket:
                _add(regions[submarket])
            if region == "ROW":
                if code not in row_countries:
                    row_countries[code] = self._empty_bucket()
                _add(row_countries[code])
                # Remember the friendliest label we saw for this code.
                row_country_labels[code] = self._label_for_country(code, market)
            _add(result["total"])

        print(f"  Matched {matched} registration_weekly rows for "
              f"week_start_mon={self.week_start_mon}")
        return self._finalize(result, row_countries, row_country_labels)

    def _finalize(self, result, row_countries, row_country_labels):
        # Sort ROW countries: highest Total first, then alphabetical.
        ordered = sorted(
            row_countries.items(),
            key=lambda kv: (-kv[1]["total_wtd"], kv[0]),
        )
        result["row_countries"] = [
            {
                "code": code,
                "label": row_country_labels.get(code, code),
                **self._stringify(bucket),
            }
            for code, bucket in ordered
        ]
        # Stringify the main region buckets and total.
        for r, bucket in result["regions"].items():
            result["regions"][r] = self._stringify(bucket)
        result["total"] = self._stringify(result["total"])
        return result


# Backwards-compatible alias.
MarketingSignupsParser = RegistrationWeeklyParser


class HTMLDashboardGenerator:
    """Generates a beautiful self-contained HTML dashboard."""

    ZAAPI_COLORS = {
        "primary": "#1e40af",      # Dark blue
        "secondary": "#0f766e",    # Dark teal
        "success": "#059669",       # Green
        "warning": "#d97706",       # Amber
        "danger": "#dc2626",        # Red
        "bg_dark": "#0f172a",       # Very dark blue
        "bg_card": "#1e293b",       # Dark slate
        "text_primary": "#f1f5f9",  # Light slate
        "text_secondary": "#cbd5e1", # Medium slate
        "border": "#334155"         # Border slate
    }

    def __init__(self):
        self.html_parts = []

    def get_target_color(self, achieved: str, target: str) -> str:
        """Determine color based on achievement percentage."""
        try:
            achieved_val = float(achieved.replace("%", "").replace(",", ""))
            if achieved_val >= 100:
                return self.ZAAPI_COLORS["success"]
            elif achieved_val >= 50:
                return self.ZAAPI_COLORS["warning"]
            else:
                return self.ZAAPI_COLORS["danger"]
        except:
            return self.ZAAPI_COLORS["text_secondary"]

    def safe_number(self, value: str, default: str = "-") -> str:
        """Safely format a number value."""
        if not value or value.strip() == "":
            return default
        return value.strip()

    def generate(self,
                 sales_data: Dict[str, Any],
                 marketing_data: Dict[str, Any],
                 output_path: str = "/tmp/dashboard.html") -> str:
        """Generate the complete HTML dashboard."""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Zaapi Daily Activity Report</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            background: linear-gradient(135deg, {self.ZAAPI_COLORS['bg_dark']} 0%, #1a2637 100%);
            color: {self.ZAAPI_COLORS['text_primary']};
            min-height: 100vh;
            padding: 20px;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
        }}

        .header {{
            background: linear-gradient(135deg, {self.ZAAPI_COLORS['primary']} 0%, {self.ZAAPI_COLORS['secondary']} 100%);
            padding: 40px 30px;
            border-radius: 12px;
            margin-bottom: 30px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.3);
        }}

        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 700;
            letter-spacing: -0.5px;
        }}

        .header .date {{
            font-size: 1.1em;
            opacity: 0.95;
            font-weight: 300;
        }}

        .summary-cards {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}

        .card {{
            background: {self.ZAAPI_COLORS['bg_card']};
            border: 1px solid {self.ZAAPI_COLORS['border']};
            border-radius: 8px;
            padding: 20px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            transition: all 0.3s ease;
        }}

        .card:hover {{
            border-color: {self.ZAAPI_COLORS['primary']};
            box-shadow: 0 8px 12px rgba(30, 64, 175, 0.15);
        }}

        .card .label {{
            font-size: 0.9em;
            color: {self.ZAAPI_COLORS['text_secondary']};
            margin-bottom: 8px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            font-weight: 600;
        }}

        .card .value {{
            font-size: 2.2em;
            font-weight: 700;
            color: {self.ZAAPI_COLORS['text_primary']};
            margin-bottom: 5px;
        }}

        .card .subtext {{
            font-size: 0.85em;
            color: {self.ZAAPI_COLORS['text_secondary']};
        }}

        .section {{
            background: {self.ZAAPI_COLORS['bg_card']};
            border: 1px solid {self.ZAAPI_COLORS['border']};
            border-radius: 8px;
            padding: 25px;
            margin-bottom: 25px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }}

        .section-title {{
            font-size: 1.5em;
            font-weight: 700;
            margin-bottom: 20px;
            padding-bottom: 10px;
            border-bottom: 2px solid {self.ZAAPI_COLORS['primary']};
            color: {self.ZAAPI_COLORS['text_primary']};
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 15px;
        }}

        thead {{
            background: rgba({int(self.ZAAPI_COLORS['primary'][1:3], 16)}, {int(self.ZAAPI_COLORS['primary'][3:5], 16)}, {int(self.ZAAPI_COLORS['primary'][5:], 16)}, 0.1);
            border-bottom: 2px solid {self.ZAAPI_COLORS['border']};
        }}

        th {{
            padding: 12px;
            text-align: left;
            font-weight: 600;
            font-size: 0.9em;
            color: {self.ZAAPI_COLORS['text_secondary']};
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        td {{
            padding: 12px;
            border-bottom: 1px solid {self.ZAAPI_COLORS['border']};
            font-size: 0.95em;
        }}

        tbody tr:hover {{
            background: rgba({int(self.ZAAPI_COLORS['primary'][1:3], 16)}, {int(self.ZAAPI_COLORS['primary'][3:5], 16)}, {int(self.ZAAPI_COLORS['primary'][5:], 16)}, 0.05);
        }}

        .metric-name {{
            font-weight: 600;
            color: {self.ZAAPI_COLORS['text_primary']};
        }}

        .metric-value {{
            font-weight: 500;
            text-align: right;
            min-width: 60px;
        }}

        .metric-target {{
            color: {self.ZAAPI_COLORS['success']};
            font-weight: 600;
        }}

        .metric-warning {{
            color: {self.ZAAPI_COLORS['warning']};
        }}

        .metric-danger {{
            color: {self.ZAAPI_COLORS['danger']};
            font-weight: 600;
        }}

        .agent-section {{
            background: rgba({int(self.ZAAPI_COLORS['primary'][1:3], 16)}, {int(self.ZAAPI_COLORS['primary'][3:5], 16)}, {int(self.ZAAPI_COLORS['primary'][5:], 16)}, 0.05);
            border-left: 4px solid {self.ZAAPI_COLORS['primary']};
            padding: 15px;
            margin-top: 15px;
            border-radius: 4px;
        }}

        .agent-name {{
            font-weight: 700;
            font-size: 1.1em;
            color: {self.ZAAPI_COLORS['primary']};
            margin-bottom: 10px;
        }}

        .account-list {{
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
            margin-top: 10px;
        }}

        .account-tag {{
            background: rgba({int(self.ZAAPI_COLORS['secondary'][1:3], 16)}, {int(self.ZAAPI_COLORS['secondary'][3:5], 16)}, {int(self.ZAAPI_COLORS['secondary'][5:], 16)}, 0.2);
            border: 1px solid {self.ZAAPI_COLORS['secondary']};
            color: {self.ZAAPI_COLORS['text_secondary']};
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 0.85em;
        }}

        .unavailable {{
            color: {self.ZAAPI_COLORS['text_secondary']};
            font-style: italic;
            padding: 40px;
            text-align: center;
            background: rgba(0, 0, 0, 0.2);
            border-radius: 8px;
        }}

        .footer {{
            text-align: center;
            padding: 20px;
            color: {self.ZAAPI_COLORS['text_secondary']};
            font-size: 0.85em;
            margin-top: 40px;
            border-top: 1px solid {self.ZAAPI_COLORS['border']};
        }}

        .footer a {{
            color: {self.ZAAPI_COLORS['primary']};
            text-decoration: none;
        }}

        .footer a:hover {{
            text-decoration: underline;
        }}

        @media (max-width: 768px) {{
            .header h1 {{
                font-size: 1.8em;
            }}

            .summary-cards {{
                grid-template-columns: 1fr;
            }}

            table {{
                font-size: 0.85em;
            }}

            th, td {{
                padding: 8px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        {self._generate_header(sales_data)}
        {self._generate_summary_cards(sales_data, marketing_data)}
        {self._generate_marketing_section(marketing_data)}
        {self._generate_sales_sections(sales_data)}
        {self._generate_renewal_section(sales_data)}
        {self._generate_intl_sales_sections(sales_data)}
        {self._generate_footer()}
    </div>
</body>
</html>
"""

        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html)
            print(f"Dashboard generated: {output_path}")
            return output_path
        except Exception as e:
            print(f"Error writing HTML file: {e}", file=sys.stderr)
            raise

    def _generate_header(self, sales_data: Dict[str, Any]) -> str:
        """Generate the header section."""
        date_str = sales_data.get("date", datetime.now().strftime("%d/%m/%Y"))
        return f"""
        <div class="header">
            <h1>Zaapi Daily Activity Report</h1>
            <div class="date">{date_str}</div>
        </div>
        """

    def _generate_summary_cards(self, sales_data: Dict[str, Any], marketing_data: Dict[str, Any]) -> str:
        """Generate summary metric cards."""

        # Extract key metrics
        outbound_won = self.safe_number(
            sales_data.get("outbound", {}).get("funnel", [{}] * 5)[-1].get("total_wtd", "-")
        )
        inbound_won = self.safe_number(
            sales_data.get("inbound", {}).get("funnel", [{}] * 5)[-1].get("total_wtd", "-")
        )

        try:
            total_won = int(float(outbound_won.replace(",", ""))) + int(float(inbound_won.replace(",", ""))) if outbound_won != "-" and inbound_won != "-" else "-"
            total_won = str(total_won) if total_won != "-" else "-"
        except:
            total_won = "-"

        outbound_activities = self.safe_number(
            sales_data.get("outbound", {}).get("funnel", [{}])[-1].get("total_wtd", "-")
        )

        marketing_total = self.safe_number(
            marketing_data.get("total", {}).get("total_wtd", "-")
        )

        return f"""
        <div class="summary-cards">
            <div class="card">
                <div class="label">Total Won Deals</div>
                <div class="value">{total_won}</div>
                <div class="subtext">Outbound + Inbound</div>
            </div>
            <div class="card">
                <div class="label">Total Activities</div>
                <div class="value">{outbound_activities}</div>
                <div class="subtext">Outbound + International Activities</div>
            </div>
            <div class="card">
                <div class="label">Total Sign-ups</div>
                <div class="value">{marketing_total}</div>
                <div class="subtext">All regions combined</div>
            </div>
            <div class="card">
                <div class="label">Dashboard Status</div>
                <div class="value">‚Äö√∫√¨</div>
                <div class="subtext">All systems operational</div>
            </div>
        </div>
        """

    def _generate_sales_sections(self, sales_data: Dict[str, Any]) -> str:
        """Generate sales outbound and inbound sections."""
        html = ""

        # Outbound Section
        html += '<div class="section">'
        sales_date = datetime.now().strftime("%d/%m/%y")
        html += f'<div class="section-title">Sales - Outbound ({sales_date})</div>'

        outbound = sales_data.get("outbound")
        if outbound and outbound.get("funnel"):
            html += self._generate_funnel_table(
                outbound.get("funnel", []),
                agents=["Yayee", "Toey"],
                agent_keys=["yayee", "toey"],
                show_targets=False
            )

            # Hot Deals section
            hot_deals = outbound.get("hot_deals", {})
            if hot_deals:
                html += '<table style="margin-top: 20px;">'
                html += '<thead><tr>'
                html += '<th>Hot Deals Category</th>'
                html += '<th style="text-align: right;">Yayee</th>'
                html += '<th style="text-align: right;">Toey</th>'
                html += '</tr></thead>'
                html += '<tbody>'

                for category in ["hot_deal", "ctp", "won"]:
                    if category in hot_deals:
                        cat_data = hot_deals[category]
                        cat_name = "Hot Deal" if category == "hot_deal" else category.upper()
                        yayee_items = cat_data.get("yayee", [])
                        toey_items = cat_data.get("toey", [])
                        yayee_str = ", ".join(yayee_items) if yayee_items else "-"
                        toey_str = ", ".join(toey_items) if toey_items else "-"
                        html += f'<tr><td class="metric-name">{cat_name}</td>'
                        html += f'<td class="metric-value">{yayee_str}</td>'
                        html += f'<td class="metric-value">{toey_str}</td>'
                        html += '</tr>'

                html += '</tbody></table>'
        else:
            html += '<div class="unavailable">Data unavailable</div>'

        html += '</div>'

        # Inbound Section
        html += '<div class="section">'
        html += f'<div class="section-title">Sales - Inbound ({sales_date})</div>'

        inbound = sales_data.get("inbound")
        if inbound and inbound.get("funnel"):
            html += self._generate_funnel_table(
                inbound.get("funnel", []),
                agents=["Pear", "Loogpad"],
                agent_keys=["pear", "loogpad"]
            )

            # Hot Deals section
            hot_deals = inbound.get("hot_deals", {})
            if hot_deals:
                html += '<table style="margin-top: 20px;">'
                html += '<thead><tr>'
                html += '<th>Hot Deals Category</th>'
                html += '<th style="text-align: right;">Pear</th>'
                html += '<th style="text-align: right;">Loogpad</th>'
                html += '</tr></thead>'
                html += '<tbody>'

                for category in ["hot_deal", "ctp", "won"]:
                    if category in hot_deals:
                        cat_data = hot_deals[category]
                        cat_name = "Hot Deal" if category == "hot_deal" else category.upper()
                        pear_items = cat_data.get("pear", [])
                        loogpad_items = cat_data.get("loogpad", [])
                        pear_str = ", ".join(pear_items) if pear_items else "-"
                        loogpad_str = ", ".join(loogpad_items) if loogpad_items else "-"
                        html += f'<tr><td class="metric-name">{cat_name}</td>'
                        html += f'<td class="metric-value">{pear_str}</td>'
                        html += f'<td class="metric-value">{loogpad_str}</td>'
                        html += '</tr>'

                html += '</tbody></table>'
        else:
            html += '<div class="unavailable">Data unavailable</div>'

        html += '</div>'

        return html

    def _generate_renewal_section(self, sales_data: Dict[str, Any]) -> str:
        """Generate the TH Account Management - Renewal section."""
        html = '<div class="section">'

        renewal = sales_data.get("renewal", {})
        header_date = renewal.get("header_date", "")
        date_label = f" ({header_date})" if header_date else ""
        html += f'<div class="section-title">TH Account Management - Renewal{date_label}</div>'

        if not renewal or (not renewal.get("due_to_renew", {}).get("pleum") and
                          not renewal.get("due_to_renew", {}).get("loogpad") and
                          not renewal.get("renewed", {}).get("pleum") and
                          not renewal.get("renewed", {}).get("loogpad")):
            html += '<div class="unavailable">Data unavailable</div>'
            html += '</div>'
            return html

        # Won metrics summary
        won = renewal.get("won", {})
        won_total = self.safe_number(won.get("total_wtd", "0"))
        won_loogpad = self.safe_number(won.get("loogpad_wtd", "0"))

        if won_total != "0" or won_loogpad != "0":
            html += f'<div style="margin-bottom: 15px; padding: 10px; background: rgba(34, 197, 94, 0.1); border-radius: 6px; border-left: 4px solid #22c55e;">'
            html += f'<span style="font-weight: 600; color: #22c55e;">Won Renewals WTD:</span> '
            html += f'Total: <strong>{won_total}</strong> | Loogpad: <strong>{won_loogpad}</strong>'
            html += '</div>'

        # Due to Renew table
        due_loogpad = renewal.get("due_to_renew", {}).get("loogpad", [])

        if due_loogpad:
            html += '<h3 style="font-size: 1.1em; margin: 15px 0 10px; color: #94a3b8;">Due to Renew This Week</h3>'

            # Loogpad column
            html += '<div>'
            html += '<div style="font-weight: 700; color: #3b82f6; margin-bottom: 8px; font-size: 1.05em;">Loogpad</div>'
            html += '<table style="width: 100%;">'
            html += '<thead><tr><th style="text-align: left; font-size: 0.8em;">Date</th>'
            html += '<th style="text-align: left; font-size: 0.8em;">Account</th>'
            html += '<th style="text-align: center; font-size: 0.8em;">Status</th></tr></thead><tbody>'
            for item in due_loogpad:
                name = item.get("name", "")
                date = item.get("date", "")
                confirmed = item.get("confirmed", False)
                if confirmed:
                    row_style = 'style="background: rgba(34, 197, 94, 0.15);"'
                    status_badge = '<span style="background: #22c55e; color: #fff; padding: 2px 8px; border-radius: 10px; font-size: 0.75em; font-weight: 600;">Confirmed</span>'
                else:
                    row_style = ''
                    status_badge = '<span style="color: #94a3b8; font-size: 0.8em;">Pending</span>'
                html += f'<tr {row_style}><td style="font-size: 0.85em; color: #94a3b8; white-space: nowrap;">{date}</td>'
                html += f'<td style="font-weight: 500;">{name}</td>'
                html += f'<td style="text-align: center;">{status_badge}</td></tr>'
            html += '</tbody></table>'
            html += '</div>'

        # Renewed section
        renewed_loogpad = renewal.get("renewed", {}).get("loogpad", [])

        if renewed_loogpad:
            html += '<h3 style="font-size: 1.1em; margin: 20px 0 10px; color: #22c55e;">&#10003; Renewed</h3>'

            html += '<div>'
            html += '<div style="font-weight: 700; color: #3b82f6; margin-bottom: 8px;">Loogpad</div>'
            html += '<div class="account-list">'
            for name in renewed_loogpad:
                html += f'<span class="account-tag" style="border-color: #22c55e; background: rgba(34, 197, 94, 0.15);">{name}</span>'
            html += '</div>'
            html += '</div>'

        # Churned section
        churned_loogpad = renewal.get("churned", {}).get("loogpad", [])

        if churned_loogpad:
            html += '<h3 style="font-size: 1.1em; margin: 20px 0 10px; color: #ef4444;">Churned</h3>'

            html += '<div>'
            html += '<div style="font-weight: 700; color: #3b82f6; margin-bottom: 8px;">Loogpad</div>'
            html += '<div class="account-list">'
            for name in churned_loogpad:
                html += f'<span class="account-tag" style="border-color: #ef4444; background: rgba(239, 68, 68, 0.15);">{name}</span>'
            html += '</div>'
            html += '</div>'

        html += '</div>'  # Close section
        return html

    def _generate_intl_sales_sections(self, sales_data: Dict[str, Any]) -> str:
        """Generate international sales sections (outbound and inbound)."""
        html = ""

        # International Outbound Section
        html += '<div class="section">'
        sales_date = datetime.now().strftime("%d/%m/%y")
        html += f'<div class="section-title">International Sales - Outbound ({sales_date})</div>'

        intl_outbound = sales_data.get("intl_outbound")
        if intl_outbound and intl_outbound.get("funnel"):
            html += self._generate_funnel_table(
                intl_outbound.get("funnel", []),
                agents=["Sheronika", "Thanom"],
                agent_keys=["sheronika", "thanom"]
            )

            # Hot Deals section
            hot_deals = intl_outbound.get("hot_deals", {})
            if hot_deals:
                html += '<table style="margin-top: 20px;">'
                html += '<thead><tr>'
                html += '<th>Hot Deals Category</th>'
                html += '<th style="text-align: right;">Sheronika</th>'
                html += '<th style="text-align: right;">Thanom</th>'
                html += '</tr></thead>'
                html += '<tbody>'

                for category in ["hot_deal", "ctp", "won"]:
                    if category in hot_deals:
                        cat_data = hot_deals[category]
                        cat_name = "Hot Deal" if category == "hot_deal" else category.upper()
                        sheronika_items = cat_data.get("sheronika", [])
                        thanom_items = cat_data.get("thanom", [])
                        sheronika_str = ", ".join(sheronika_items) if sheronika_items else "-"
                        thanom_str = ", ".join(thanom_items) if thanom_items else "-"
                        html += f'<tr><td class="metric-name">{cat_name}</td>'
                        html += f'<td class="metric-value">{sheronika_str}</td>'
                        html += f'<td class="metric-value">{thanom_str}</td>'
                        html += '</tr>'

                html += '</tbody></table>'
        else:
            html += '<div class="unavailable">Data unavailable</div>'

        html += '</div>'

        # International Inbound Section
        html += '<div class="section">'
        html += f'<div class="section-title">International Sales - Inbound ({sales_date})</div>'

        intl_inbound = sales_data.get("intl_inbound")
        if intl_inbound and intl_inbound.get("funnel"):
            html += self._generate_funnel_table(
                intl_inbound.get("funnel", []),
                agents=["Sheronika", "Thanom"],
                agent_keys=["sheronika", "thanom"]
            )

            # Hot Deals section
            hot_deals = intl_inbound.get("hot_deals", {})
            if hot_deals:
                html += '<table style="margin-top: 20px;">'
                html += '<thead><tr>'
                html += '<th>Hot Deals Category</th>'
                html += '<th style="text-align: right;">Sheronika</th>'
                html += '<th style="text-align: right;">Thanom</th>'
                html += '</tr></thead>'
                html += '<tbody>'

                for category in ["hot_deal", "ctp", "won"]:
                    if category in hot_deals:
                        cat_data = hot_deals[category]
                        cat_name = "Hot Deal" if category == "hot_deal" else category.upper()
                        sheronika_items = cat_data.get("sheronika", [])
                        thanom_items = cat_data.get("thanom", [])
                        sheronika_str = ", ".join(sheronika_items) if sheronika_items else "-"
                        thanom_str = ", ".join(thanom_items) if thanom_items else "-"
                        html += f'<tr><td class="metric-name">{cat_name}</td>'
                        html += f'<td class="metric-value">{sheronika_str}</td>'
                        html += f'<td class="metric-value">{thanom_str}</td>'
                        html += '</tr>'

                html += '</tbody></table>'
        else:
            html += '<div class="unavailable">Data unavailable</div>'

        html += '</div>'

        return html

    def _generate_funnel_table(self, funnel: List[Dict[str, str]],
                               agents: List[str], agent_keys: List[str],
                               show_targets: bool = True) -> str:
        """Generate a funnel metrics table."""
        html = '<table>'
        html += '<thead><tr>'
        html += '<th>Metric</th>'
        html += '<th style="text-align: right;">Total WTD</th>'
        html += '<th style="text-align: right;">Total Daily</th>'
        if show_targets:
            html += '<th style="text-align: right;">WTD vs Target</th>'

        for agent in agents:
            html += f'<th style="text-align: right;">{agent} WTD</th>'
            html += f'<th style="text-align: right;">{agent} Daily</th>'
            if show_targets:
                html += f'<th style="text-align: right;">{agent} vs Target</th>'

        html += '</tr></thead><tbody>'

        for metric in funnel:
            metric_name = metric.get("name", "")
            html += f'<tr><td class="metric-name">{metric_name}</td>'

            # Total columns
            total_wtd = self.safe_number(metric.get("total_wtd", "-"))
            total_daily = self.safe_number(metric.get("total_daily", "-"))
            total_vs_target = self.safe_number(metric.get("total_vs_target", "-"))

            html += f'<td class="metric-value">{total_wtd}</td>'
            html += f'<td class="metric-value">{total_daily}</td>'

            if show_targets:
                # Target color
                color = self.get_target_color(total_vs_target, "100%")
                target_class = "metric-target" if color == self.ZAAPI_COLORS["success"] else \
                              "metric-warning" if color == self.ZAAPI_COLORS["warning"] else "metric-danger"
                html += f'<td class="metric-value {target_class}">{total_vs_target}</td>'

            # Agent columns
            for agent_key in agent_keys:
                wtd = self.safe_number(metric.get(f"{agent_key}_wtd", "-"))
                daily = self.safe_number(metric.get(f"{agent_key}_daily", "-"))
                vs_target = self.safe_number(metric.get(f"{agent_key}_vs_target", "-"))

                html += f'<td class="metric-value">{wtd}</td>'
                html += f'<td class="metric-value">{daily}</td>'

                if show_targets:
                    color = self.get_target_color(vs_target, "100%")
                    target_class = "metric-target" if color == self.ZAAPI_COLORS["success"] else \
                                  "metric-warning" if color == self.ZAAPI_COLORS["warning"] else "metric-danger"
                    html += f'<td class="metric-value {target_class}">{vs_target}</td>'

            html += '</tr>'

        html += '</tbody></table>'
        return html

    def _generate_marketing_section(self, marketing_data: Dict[str, Any]) -> str:
        """Generate the Marketing - Lead Overview section.

        Columns: Region | Qualified | HQ+ | Total | Attributed | Unattributed
        - Total = Qualified + HQ+
        - Attributed + Unattributed = Total
        Rows (in order):
          GLOBAL (grand total),
          TH,
          SEA, ‚Äö√ú‚â• MY, ‚Äö√ú‚â• SG, ‚Äö√ú‚â• PH,
          ROW, ‚Äö√ú‚â• <country code> for every ROW country with data this week.
        Source: registration_weekly tab, filtered to current ISO week (WTD).
        """
        html = '<div class="section">'
        marketing_date = datetime.now().strftime("%d/%m/%y")
        html += (
            f'<div class="section-title">Marketing - Lead Overview '
            f'({marketing_date}, WTD)</div>'
        )

        regions_data = marketing_data.get("regions", {})
        row_countries = marketing_data.get("row_countries", [])

        if not regions_data:
            html += '<div class="unavailable">Data unavailable</div>'
            html += '</div>'
            return html

        html += '<table>'
        html += '<thead><tr>'
        html += '<th>Region</th>'
        html += '<th style="text-align: right;">Qualified</th>'
        html += '<th style="text-align: right;">HQ+</th>'
        html += '<th style="text-align: right;">Total</th>'
        html += '<th style="text-align: right;">Attributed</th>'
        html += '<th style="text-align: right;">Unattributed</th>'
        html += '</tr></thead><tbody>'

        def render_row(label, data, indent=False, emphasize=False):
            qualified = self.safe_number(data.get("qualified_wtd", "-"))
            hq = self.safe_number(data.get("hqplus_wtd", "-"))
            total = self.safe_number(data.get("total_wtd", "-"))
            attributed = self.safe_number(data.get("attributed_wtd", "-"))
            unattributed = self.safe_number(data.get("unattributed_wtd", "-"))

            name_style = ' style="padding-left: 28px; opacity: 0.85;"' if indent else ''
            tr_style = ''
            if emphasize:
                # Bold the GLOBAL row so the grand total reads as the header line.
                tr_style = ' style="font-weight: 700; border-top: 2px solid #334155; border-bottom: 2px solid #334155;"'
            display = ("&nbsp;&nbsp;&nbsp;&nbsp;‚Äö√ú‚â• " + label) if indent else label

            row_html = f'<tr{tr_style}><td class="metric-name"{name_style}>{display}</td>'
            row_html += f'<td class="metric-value">{qualified}</td>'
            row_html += f'<td class="metric-value">{hq}</td>'
            row_html += f'<td class="metric-value">{total}</td>'
            row_html += f'<td class="metric-value">{attributed}</td>'
            row_html += f'<td class="metric-value">{unattributed}</td>'
            row_html += '</tr>'
            return row_html

        # GLOBAL grand total
        html += render_row("GLOBAL",
                           regions_data.get("GLOBAL", {}),
                           indent=False, emphasize=True)
        # TH
        html += render_row("TH", regions_data.get("TH", {}))
        # SEA + submarkets
        html += render_row("SEA", regions_data.get("SEA", {}))
        for sub in ("MY", "SG", "PH"):
            html += render_row(sub, regions_data.get(sub, {}), indent=True)
        # ROW + country submarkets
        html += render_row("ROW", regions_data.get("ROW", {}))
        for c in row_countries:
            label = f"{c.get('code', '??')} ({c.get('label', '')})" \
                    if c.get("label") and c.get("label") != c.get("code") \
                    else c.get("code", "??")
            html += render_row(label, c, indent=True)

        html += '</tbody></table>'
        html += '</div>'
        return html

    def _generate_footer(self) -> str:
        """Generate the footer section."""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return f"""
        <div class="footer">
            <p>Last updated: {now}</p>
            <p>Zaapi Daily Activity Report - Automatically Generated</p>
        </div>
        """


class SlackNotifier:
    """Handles posting to Slack via Incoming Webhook."""

    def __init__(self, webhook_url: str):
        self.webhook_url = webhook_url

    def post_summary(self, sales_data: Dict[str, Any], marketing_data: Dict[str, Any], dashboard_url: str) -> bool:
        """Post a summary message to Slack."""
        try:
            date_str = sales_data.get("date", datetime.now().strftime("%d/%m/%Y"))

            # Extract sales metrics
            outbound_funnel = sales_data.get("outbound", {}).get("funnel", [])
            inbound_funnel = sales_data.get("inbound", {}).get("funnel", [])
            intl_outbound_funnel = sales_data.get("intl_outbound", {}).get("funnel", [])
            intl_inbound_funnel = sales_data.get("intl_inbound", {}).get("funnel", [])

            # Find Won metric by name (international sections have extra rows after Won)
            def find_won_wtd(funnel):
                for m in funnel:
                    if m.get("name", "").lower().strip() == "won":
                        val = m.get("total_wtd", "0")
                        return val if val and val != "-" else "0"
                # Fallback: last metric
                return funnel[-1].get("total_wtd", "0") if funnel else "0"

            outbound_won = find_won_wtd(outbound_funnel)
            inbound_won = find_won_wtd(inbound_funnel)
            intl_out_won = find_won_wtd(intl_outbound_funnel)
            intl_in_won = find_won_wtd(intl_inbound_funnel)

            outbound_contact = outbound_funnel[1].get("total_wtd", "0") if len(outbound_funnel) > 1 else "0"
            inbound_contact = inbound_funnel[1].get("total_wtd", "0") if len(inbound_funnel) > 1 else "0"

            # Extract marketing metrics (registration_weekly aggregates)
            mktg_total = marketing_data.get("total", {})
            mktg_regions = marketing_data.get("regions", {})
            total_leads_wtd = mktg_total.get("total_wtd", "0")
            qualified_wtd = mktg_total.get("qualified_wtd", "0")
            hqplus_wtd = mktg_total.get("hqplus_wtd", "0")

            def _mkt_row(label, key):
                d = mktg_regions.get(key, {})
                q = d.get("qualified_wtd", "0")
                h = d.get("hqplus_wtd", "0")
                t = d.get("total_wtd", "0")
                # Fixed-width row inside back-ticks so columns line up in Slack.
                return f"` {label:<6} {q:>3}   {h:>3}    {t:>3} `"

            marketing_table = "\n".join([
                "` Region    Q   HQ+  Total `",
                _mkt_row("TH",  "TH"),
                _mkt_row("SEA", "SEA"),
                _mkt_row("ROW", "ROW"),
            ])

            message = f"""
:chart_with_upwards_trend: *Zaapi Daily Activity Report ‚Äö√Ñ√Æ {date_str}*

*Marketing ‚Äö√Ñ√Æ Lead Overview (WTD)*
*GLOBAL:* {total_leads_wtd} total  |  Qualified: {qualified_wtd}  |  HQ+: {hqplus_wtd}
{marketing_table}

*Sales ‚Äö√Ñ√Æ Won Deals WTD*
‚Äö√Ñ¬¢ Outbound: *{outbound_won}*  |  Inbound: *{inbound_won}*
‚Äö√Ñ¬¢ Intl Outbound: *{intl_out_won}*  |  Intl Inbound: *{intl_in_won}*

*Sales ‚Äö√Ñ√Æ Contacts WTD*
‚Äö√Ñ¬¢ Outbound: *{outbound_contact}*  |  Inbound: *{inbound_contact}*

:link: <{dashboard_url}|View Full Dashboard>
            """.strip()

            payload = {
                "text": message
            }

            response = requests.post(
                self.webhook_url,
                json=payload,
                timeout=10
            )

            response.raise_for_status()

            if response.text == "ok":
                print("Slack message posted successfully via webhook")
                return True
            else:
                print(f"Slack webhook error: {response.text}", file=sys.stderr)
                return False

        except Exception as e:
            print(f"Error posting to Slack: {e}", file=sys.stderr)
            return False


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Fetch Google Sheets data and generate Zaapi Daily Activity Report"
    )
    parser.add_argument(
        "--output",
        default="/tmp/zaapi_report.html",
        help="Output HTML file path"
    )
    parser.add_argument(
        "--slack-webhook-url",
        help="Slack Incoming Webhook URL for posting notifications"
    )
    parser.add_argument(
        "--github-pages-url",
        help="URL to GitHub Pages hosted dashboard"
    )

    args = parser.parse_args()

    print("Starting Zaapi Daily Activity Report generation...")

    # Fetch data
    fetcher = GoogleSheetsFetcher()

    print("Fetching Sales Huddle data...")
    week_name = fetcher.get_current_week_sheet_name()
    print(f"Current week: {week_name}")

    # Fetch XLSX first for sheet name resolution and green cell detection
    sales_sheet_id = "1A33NpnkZlgrwyDSOKn3nwB0u5mFtal2BlVC0nGZL7Xk"
    print("Fetching XLSX for sheet name resolution and renewal formatting...")
    wb = fetcher.fetch_sheet_xlsx(sales_sheet_id)
    week_name = fetcher.resolve_sheet_name_from_wb(wb, week_name)

    sales_sheet = fetcher.fetch_sheet(
        sales_sheet_id,
        week_name
    )

    # Marketing leads now sourced from the Zaapi-growth Ads Data sheet
    # (tab: registration_weekly). We aggregate the current ISO week's rows by
    # market into TH / SEA / MY / SG / PH / ROW buckets ‚Äö√Ñ√Æ see
    # RegistrationWeeklyParser.
    print("Fetching Marketing Lead Overview (registration_weekly)...")
    marketing_sheet = fetcher.fetch_sheet(
        "1s5AC58mAylpSDknU7L7HRJUPrVf36b0TvzD35tW-Wdw",
        "registration_weekly",
        gid="859536577",
    )
    # ISO Monday for the current week, in YYYY-MM-DD (matches the
    # week_start_mon column format in registration_weekly).
    _today = datetime.now().date()
    current_week_start_mon = (_today - timedelta(days=_today.weekday())).isoformat()
    print(f"  Current week_start_mon: {current_week_start_mon}")

    # Parse data
    sales_data = {
        "date": datetime.now().strftime("%d/%m/%Y"),
        "outbound": {},
        "inbound": {},
        "renewal": {},
        "intl_outbound": {},
        "intl_inbound": {}
    }

    if sales_sheet:
        # Fetch hot deals sections with range-specific queries to work around
        # Google Sheets merged cell CSV export issues that drop agent data
        print("Fetching Hot Deals ranges (merged cell workaround)...")
        hot_deals_ranges = {
            "outbound": fetcher.fetch_sheet(sales_sheet_id, week_name, cell_range="J19:O55"),
            "inbound": fetcher.fetch_sheet(sales_sheet_id, week_name, cell_range="AA19:AI55"),
            "intl_inbound": fetcher.fetch_sheet(sales_sheet_id, week_name, cell_range="AU19:AZ55"),
            "intl_outbound": fetcher.fetch_sheet(sales_sheet_id, week_name, cell_range="BK19:BP55"),
        }

        # Fetch renewal section data with range-specific query
        print("Fetching Renewal section data...")
        renewal_range = fetcher.fetch_sheet(sales_sheet_id, week_name, cell_range="R58:AF100")

        # Use already-fetched XLSX for green cell detection (confirmed renewals)
        print("Checking XLSX for renewal formatting...")
        green_cells = set()
        try:
            if wb and week_name in wb.sheetnames:
                ws = wb[week_name]
                # Check account name columns (AB=28, AE=31) for green background
                # Due to renew entries start at row 61 in the actual sheet
                for row in range(61, 100):
                    for col in [28, 31]:  # AB and AE columns
                        cell = ws.cell(row=row, column=col)
                        name = str(cell.value).strip() if cell.value else ""
                        if name and cell.fill and cell.fill.fgColor:
                            rgb = str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb else ""
                            if len(rgb) >= 6:
                                try:
                                    # Handle both 6-char and 8-char RGB (with alpha prefix)
                                    if len(rgb) == 8:
                                        r_val = int(rgb[2:4], 16)
                                        g_val = int(rgb[4:6], 16)
                                        b_val = int(rgb[6:8], 16)
                                    else:
                                        r_val = int(rgb[0:2], 16)
                                        g_val = int(rgb[2:4], 16)
                                        b_val = int(rgb[4:6], 16)
                                    # Green if green component is dominant
                                    if g_val > r_val and g_val > b_val and g_val > 100:
                                        green_cells.add(name)
                                except (ValueError, IndexError):
                                    pass
                wb.close()
        except Exception as e:
            print(f"Warning: Could not detect green cells: {e}", file=sys.stderr)

        if green_cells:
            print(f"Found {len(green_cells)} confirmed renewal(s): {green_cells}")

        parser = SalesHuddleParser(sales_sheet, hot_deals_ranges=hot_deals_ranges)
        date_info = parser.get_date_info()
        sales_data["date"] = date_info.get("date", sales_data["date"])
        sales_data["outbound"] = parser.parse_outbound_section()
        sales_data["inbound"] = parser.parse_inbound_section()
        sales_data["renewal"] = parser.parse_renewal_section(
            renewal_range_data=renewal_range,
            green_cells=green_cells
        )
        sales_data["intl_outbound"] = parser.parse_intl_outbound_section()
        sales_data["intl_inbound"] = parser.parse_intl_inbound_section()
    else:
        print("Warning: Sales Huddle data not available", file=sys.stderr)

    marketing_data = {
        "date": datetime.now().strftime("%d/%m/%Y"),
        "regions": {},
        "total": {}
    }

    if marketing_sheet:
        parser = RegistrationWeeklyParser(marketing_sheet, current_week_start_mon)
        marketing_data = parser.parse_data()
    else:
        print("Warning: registration_weekly data not available", file=sys.stderr)

    # Generate HTML
    generator = HTMLDashboardGenerator()
    output_path = generator.generate(sales_data, marketing_data, args.output)

    print(f"Dashboard generated successfully: {output_path}")

    # Post to Slack if configured
    if args.slack_webhook_url:
        notifier = SlackNotifier(args.slack_webhook_url)
        dashboard_url = args.github_pages_url or output_path
        notifier.post_summary(sales_data, marketing_data, dashboard_url)

    print("Report generation complete!")


if __name__ == "__main__":
    main()
